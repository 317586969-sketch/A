# -*- coding: utf-8 -*-
"""
AnomalyDetect_new.py — v4.x
TireMonitor Anomaly Detection (007)
"""
from __future__ import annotations
import os, time, traceback, threading
from datetime import datetime
from typing import Dict, List
from concurrent.futures import ThreadPoolExecutor, as_completed
import pandas as pd
import pyodbc
import datarobot as dr
from datarobot_predict.deployment import predict
import config
import config_anomaly

# ============ 1. Config ============
SQL_CONN_STR = config.SQL_CONFIG
DATAROBOT_API_TOKEN = config.API_TOKEN
VIEW_NAME = config_anomaly.ANOMALY_PENDING_VIEW
STATUS_TABLE = config_anomaly.ANOMALY_STATUS_TABLE
ANOMALY_HISTORICAL_FEATURE_VIEW = config_anomaly.ANOMALY_HISTORICAL_FEATURE_VIEW
BATCH_SIZE = config_anomaly.BATCH_SIZE
DEPLOYMENT_CONFIG = config_anomaly.DEPLOYMENT_CONFIG
MAX_EXPLANATIONS = config_anomaly.MAX_EXPLANATIONS
ANOMALY_UCL = getattr(config_anomaly, "ANOMALY_UCL", 0.85)
ANOMALY_UWL = getattr(config_anomaly, "ANOMALY_UWL", 0.70)
ANOMALY_RULE2_WINDOW = getattr(config_anomaly, "ANOMALY_RULE2_WINDOW", 5)
ANOMALY_SPC_SIGMA = getattr(config_anomaly, "ANOMALY_SPC_SIGMA", 3.0)
ANOMALY_SPC_MIN_STD = getattr(config_anomaly, "ANOMALY_SPC_MIN_STD", 0.5)
ENABLE_SPC_RULE = getattr(config_anomaly, "ENABLE_SPC_RULE", False)
ENABLE_EXCEL_REPORT = getattr(config_anomaly, "ENABLE_EXCEL_REPORT", False)
ENABLE_AI_DIAGNOSIS = getattr(config_anomaly, "ENABLE_AI_DIAGNOSIS", False)
GEN_AI_URL = getattr(config, "GEN_AI_URL", "")
GEN_AI_TOKEN = getattr(config, "GEN_AI_TOKEN", DATAROBOT_API_TOKEN)
DISABLED_ANOMALY_FEATURES = set(getattr(config_anomaly, "DISABLED_ANOMALY_FEATURES", set()))
EXCLUDE_DISABLED_FEATURES_FROM_PREDICTION = getattr(config_anomaly, "EXCLUDE_DISABLED_FEATURES_FROM_PREDICTION", False)
NULL_DISABLED_FEATURES_FOR_PREDICTION = getattr(config_anomaly, "NULL_DISABLED_FEATURES_FOR_PREDICTION", False)
ENABLE_SENSOR_FAULT_SUPPRESSION = getattr(config_anomaly, "ENABLE_SENSOR_FAULT_SUPPRESSION", True)
EXCLUDE_COLUMNS = {"Band_drum_Drum_No_Act", "Belt_drum_Drum_No_Act", "ROUTE_KEY"}
LOG_LOCK = threading.Lock()
os.makedirs(config.LOG_DIR, exist_ok=True)
LOG_FILE = os.path.join(config.LOG_DIR, f"Log_AnomalyDetect_{datetime.now().strftime('%Y%m%d')}.txt")
ANOMALY_MAIL_ENABLE = getattr(config_anomaly, "ANOMALY_MAIL_ENABLE", 1)
ANOMALY_MAIL_TO = getattr(config_anomaly, "ANOMALY_MAIL_TO", "")
ANOMALY_MAIL_CC = getattr(config_anomaly, "ANOMALY_MAIL_CC", "")
ENABLE_EXPLANATION_TREND_GRAPH = getattr(config_anomaly, "ENABLE_EXPLANATION_TREND_GRAPH", True)
EXPLANATION_TREND_WINDOW_MINUTES_BEFORE = getattr(config_anomaly, "EXPLANATION_TREND_WINDOW_MINUTES_BEFORE", 120)
EXPLANATION_TREND_WINDOW_MINUTES_AFTER = getattr(config_anomaly, "EXPLANATION_TREND_WINDOW_MINUTES_AFTER", 120)
EXPLANATION_TREND_TARGET_STATUSES = getattr(config_anomaly, "EXPLANATION_TREND_TARGET_STATUSES", ["ANOMALY", "TREND_ANOMALY", "WARNING", "SENSOR_FAULT_SUPPRESSED"])
EXPLANATION_TREND_MAX_EVENTS = getattr(config_anomaly, "EXPLANATION_TREND_MAX_EVENTS", 20)
EXPLANATION_TREND_OUTPUT_SUBDIR = getattr(config_anomaly, "EXPLANATION_TREND_OUTPUT_SUBDIR", "explanation_trend_graphs")

print("[CONFIG] AnomalyDetect v4.x initialized successfully")
print("[CONFIG] All module-level config loaded — syntax OK")

# ============ 2. Logging ============
def log_message(message: str) -> None:
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{timestamp}] {message}"
    print(line)
    try:
        with LOG_LOCK:
            with open(LOG_FILE, "a", encoding="utf-8") as f_log:
                f_log.write(line + "\n")
    except Exception as e:
        print(f"[LOGGER_ERROR] {e}")

def log_elapsed(message: str, start_time: float) -> None:
    elapsed = time.perf_counter() - start_time
    log_message(f"{message} elapsed={elapsed:.2f}s")

# ============ 3. SQL Utility ============
def get_connection() -> pyodbc.Connection:
    return pyodbc.connect(SQL_CONN_STR)

def fetch_pending_data() -> pd.DataFrame:
    t0 = time.perf_counter()
    sql = f"SELECT TOP 100 * FROM {VIEW_NAME} ORDER BY Production_DateTime DESC"
    log_message("[SQL] start fetch pending data")
    with get_connection() as conn:
        df = pd.read_sql(sql, conn)
    log_elapsed(f"[SQL] finish fetch pending data rows={len(df)}", t0)
    return df

def insert_status_rows(rows: List[dict]) -> None:
    if not rows:
        return
    t0 = time.perf_counter()
    insert_sql = f"""
    INSERT INTO {STATUS_TABLE} (
        Machine_No, Product_No_TBM, Production_DateTime, Serial_No_TBM,
        status, message, anomaly_score, judgment, trigger_rule,
        deployment_key, route_key, top_explanation, processed_at
    )
    SELECT ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, SYSDATETIME()
    WHERE NOT EXISTS (
        SELECT 1 FROM {STATUS_TABLE}
        WHERE Machine_No = ? AND Product_No_TBM = ? AND Production_DateTime = ?
          AND Serial_No_TBM = ? AND route_key = ?
    )
    """
    insert_params = []
    for r in rows:
        params = (
            r.get("Machine_No"), r.get("Product_No_TBM"), r.get("Production_DateTime"),
            r.get("Serial_No_TBM"), r.get("status"), r.get("message"),
            r.get("anomaly_score"), r.get("judgment"), r.get("trigger_rule"),
            r.get("deployment_key"), r.get("route_key"), r.get("top_explanation"),
            r.get("Machine_No"), r.get("Product_No_TBM"), r.get("Production_DateTime"),
            r.get("Serial_No_TBM"), r.get("route_key"),
        )
        insert_params.append(params)
    try:
        with get_connection() as conn:
            cur = conn.cursor()
            cur.fast_executemany = True
            cur.executemany(insert_sql, insert_params)
            conn.commit()
            log_elapsed(f"[SQL] insert status rows attempted={len(rows)}", t0)
    except Exception as e:
        log_message(f"[SQL] insert_status_rows failed:\n" + traceback.format_exc())
        raise

print("[SQL] utility functions loaded")

# ============ 4. Routing ============
def validate_required_columns(df: pd.DataFrame) -> None:
    required = ["Machine_No", "Product_No_TBM", "Production_DateTime", "Serial_No_TBM", "ROUTE_KEY"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Missing columns: {missing}")

def split_by_routing(df: pd.DataFrame) -> Dict[tuple, pd.DataFrame]:
    validate_required_columns(df)
    df = df.copy()
    df["Machine_No_key"] = pd.to_numeric(df["Machine_No"], errors="coerce").astype("Int64")
    valid_route_keys = {route_key for (_, route_key) in DEPLOYMENT_CONFIG.keys()}
    invalid_df = df[df["Machine_No_key"].isna() | ~df["ROUTE_KEY"].isin(valid_route_keys)].copy()
    if not invalid_df.empty:
        invalid_rows = []
        for _, row in invalid_df.iterrows():
            invalid_rows.append({
                "Machine_No": row.get("Machine_No"), "Product_No_TBM": row.get("Product_No_TBM"),
                "Production_DateTime": row.get("Production_DateTime"), "Serial_No_TBM": row.get("Serial_No_TBM"),
                "status": "SKIPPED_INVALID_ROUTE", "judgment": "SKIPPED_INVALID_ROUTE",
                "trigger_rule": "INVALID_ROUTE", "route_key": row.get("ROUTE_KEY"),
            })
        insert_status_rows(invalid_rows)
    valid_df = df[df["Machine_No_key"].notna() & df["ROUTE_KEY"].isin(valid_route_keys)].copy()
    routed: Dict[tuple, pd.DataFrame] = {}
    skipped_no_model_rows = []
    for (machine_no, route_key), sub_df in valid_df.groupby(["Machine_No_key", "ROUTE_KEY"], dropna=False):
        model_key = (int(machine_no), route_key)
        if model_key not in DEPLOYMENT_CONFIG:
            for _, row in sub_df.iterrows():
                skipped_no_model_rows.append({
                    "Machine_No": row.get("Machine_No"), "Product_No_TBM": row.get("Product_No_TBM"),
                    "Production_DateTime": row.get("Production_DateTime"), "Serial_No_TBM": row.get("Serial_No_TBM"),
                    "status": "SKIPPED_NO_MODEL", "judgment": "SKIPPED_NO_MODEL",
                    "trigger_rule": "NO_DEPLOYMENT_MODEL", "route_key": row.get("ROUTE_KEY"),
                })
            continue
        sub_df = sub_df.drop(columns=["Machine_No_key"])
        routed[model_key] = sub_df.copy()
    if skipped_no_model_rows:
        insert_status_rows(skipped_no_model_rows)
    return routed

print("[ROUTING] loaded")

# ============ 4.5 Anomaly Judgment ============
def normalize_feature_name(name):
    if name is None: return ""
    return str(name).strip().replace("-", "_")

def is_disabled_feature(feature_name):
    target = normalize_feature_name(feature_name)
    disabled_normalized = {normalize_feature_name(x) for x in DISABLED_ANOMALY_FEATURES}
    return target in disabled_normalized

def is_prediction_excluded_column(col_name):
    if col_name in EXCLUDE_COLUMNS: return True
    if not EXCLUDE_DISABLED_FEATURES_FROM_PREDICTION: return False
    return is_disabled_feature(col_name)

def find_anomaly_score_column(df):
    for col in ["ANOMALY_SCORE", "anomaly_score", "Anomaly Score", "prediction", "Prediction"]:
        if col in df.columns: return col
    raise ValueError(f"anomaly score column not found: {list(df.columns)}")

def judge_by_fixed_threshold(score):
    if pd.isna(score): return "ERROR_DATA", "SCORE_NULL"
    score = float(score)
    if score >= ANOMALY_UCL: return "ANOMALY", "RULE1_UCL"
    if score >= ANOMALY_UWL: return "WARNING", "RULE0_UWL"
    return "NORMAL", "RULE0_NORMAL"

def get_explanation_feature_columns(row):
    cols = [c for c in row.index if "EXPLANATION" in str(c).upper() and "FEATURE_NAME" in str(c).upper()]
    def sort_key(col):
        parts = str(col).split("_")
        for p in parts:
            if p.isdigit(): return int(p)
        return 999
    return sorted(cols, key=sort_key)

def extract_first_explanation_feature(row):
    for col in get_explanation_feature_columns(row):
        value = row.get(col)
        if pd.notna(value) and str(value).strip(): return str(value).strip()[:500]
    return None

def extract_top_explanation(row):
    for col in get_explanation_feature_columns(row):
        value = row.get(col)
        if pd.isna(value): continue
        feature_name = str(value).strip()
        if not feature_name: continue
        if is_disabled_feature(feature_name): continue
        return feature_name[:500]
    return None

def apply_sensor_fault_suppression(df):
    if not ENABLE_SENSOR_FAULT_SUPPRESSION: return df
    df = df.copy()
    df["top_explanation_raw"] = df.apply(extract_first_explanation_feature, axis=1)
    df["suppressed_feature"] = df["top_explanation_raw"].apply(lambda x: x if is_disabled_feature(x) else None)
    suppress_mask = df["suppressed_feature"].notna() & df["judgment"].isin(["WARNING", "ANOMALY", "TREND_ANOMALY"])
    df.loc[suppress_mask, "judgment"] = "SENSOR_FAULT_SUPPRESSED"
    df.loc[suppress_mask, "trigger_rule"] = "DISABLED_FEATURE_TOP_EXPLANATION"
    return df

def add_judgment_columns(pred_df):
    df = pred_df.copy()
    score_col = find_anomaly_score_column(df)
    df["anomaly_score"] = pd.to_numeric(df[score_col], errors="coerce")
    judgments = df["anomaly_score"].apply(judge_by_fixed_threshold)
    df["judgment"] = judgments.apply(lambda x: x[0])
    df["trigger_rule"] = judgments.apply(lambda x: x[1])
    df["top_explanation"] = df.apply(extract_top_explanation, axis=1)
    df = apply_sensor_fault_suppression(df)
    return df

def _fetch_historical_scores(machine_no, route_key, limit):
    if machine_no is None or pd.isna(machine_no) or route_key is None or pd.isna(route_key):
        return pd.DataFrame()
    sql = f"SELECT TOP {int(limit)} anomaly_score, Production_DateTime FROM {STATUS_TABLE} WHERE Machine_No = ? AND route_key = ? AND anomaly_score IS NOT NULL ORDER BY Production_DateTime DESC"
    try:
        with get_connection() as conn:
            df = pd.read_sql(sql, conn, params=[machine_no, route_key])
        if df.empty: return pd.DataFrame()
        return df.sort_values("Production_DateTime").reset_index(drop=True)
    except Exception:
        return pd.DataFrame()

def apply_rule2_consecutive_uwl(pred_df):
    if pred_df.empty: return pred_df
    required_cols = ["Machine_No", "Product_No_TBM", "ROUTE_KEY", "Production_DateTime", "anomaly_score", "judgment", "trigger_rule"]
    missing = [c for c in required_cols if c not in pred_df.columns]
    if missing:
        log_message(f"[RULE2] skipped. missing columns={missing}")
        return pred_df
    df = pred_df.copy()
    df["_rule2_datetime"] = pd.to_datetime(df["Production_DateTime"], errors="coerce")
    df["_rule2_score"] = pd.to_numeric(df["anomaly_score"], errors="coerce")
    df["_rule2_idx"] = df.index
    total_trend = 0
    for (machine_no, route_key), grp in df.groupby(["Machine_No", "ROUTE_KEY"], dropna=False):
        if pd.isna(machine_no) or pd.isna(route_key): continue
        hist_limit = ANOMALY_RULE2_WINDOW - 1
        hist_df = _fetch_historical_scores(machine_no, route_key, hist_limit)
        if hist_df.empty:
            hist_scores = []
        else:
            hist_scores = pd.to_numeric(hist_df["anomaly_score"], errors="coerce").tolist()
        grp_sorted = grp.sort_values("_rule2_datetime")
        batch_scores = grp_sorted["_rule2_score"].tolist()
        combined_scores = hist_scores + batch_scores
        hist_len = len(hist_scores)
        consecutive = 0
        for i, score in enumerate(combined_scores):
            if pd.notna(score) and float(score) >= ANOMALY_UWL:
                consecutive += 1
            else:
                consecutive = 0
            if i >= hist_len:
                batch_pos = i - hist_len
                row_idx = grp_sorted.iloc[batch_pos]["_rule2_idx"]
                if consecutive >= ANOMALY_RULE2_WINDOW:
                    current_judgment = str(df.at[row_idx, "judgment"])
                    if current_judgment != "ANOMALY":
                        df.at[row_idx, "judgment"] = "TREND_ANOMALY"
                        df.at[row_idx, "trigger_rule"] = f"RULE2_CONSECUTIVE_UWL_{ANOMALY_RULE2_WINDOW}"
                        total_trend += 1
    if total_trend > 0:
        log_message(f"[RULE2] consecutive UWL detected (cross-batch). TREND_ANOMALY={total_trend}")
    df = df.drop(columns=["_rule2_datetime", "_rule2_score", "_rule2_idx"], errors="ignore")
    return df

print("[JUDGMENT] loaded")

# ============ 5. DataRobot Prediction ============
def call_datarobot_api(model_key, df_part):
    t0 = time.perf_counter()
    deployment_info = DEPLOYMENT_CONFIG[model_key]
    deployment_id = deployment_info["deployment_id"]
    deployment_key = deployment_info["deployment_key"]
    log_message(f"[PREDICT] start {deployment_key} Machine_No={model_key[0]} ROUTE_KEY={model_key[1]} rows={len(df_part)}")
    dr.Client(endpoint=getattr(config, 'ENDPOINT', "https://app.datarobot.com/api/v2"), token=DATAROBOT_API_TOKEN, trace_context="AnomalyDetectRealtime")
    send_cols = [c for c in df_part.columns if not is_prediction_excluded_column(c)]
    df_send = df_part[send_cols].copy()
    if NULL_DISABLED_FEATURES_FOR_PREDICTION:
        for col in df_send.columns:
            if is_disabled_feature(col):
                df_send[col] = None
    for col in df_send.columns:
        if pd.api.types.is_datetime64_any_dtype(df_send[col]):
            df_send[col] = df_send[col].astype(str)
    deployment = dr.Deployment.get(deployment_id)
    pred_df, _ = predict(deployment=deployment, data_frame=df_send, max_explanations=MAX_EXPLANATIONS)
    key_cols = ["Machine_No", "Product_No_TBM", "Production_DateTime", "Serial_No_TBM", "ROUTE_KEY"]
    pred_df = pd.concat([df_part[key_cols].reset_index(drop=True), pred_df.reset_index(drop=True)], axis=1)
    pred_df["deployment_key"] = deployment_key
    pred_df = add_judgment_columns(pred_df)
    log_elapsed(f"[PREDICT] finish {deployment_key} rows={len(df_part)}", t0)
    return pred_df

def classify_prediction_exception(error):
    error_text = f"{type(error).__name__}: {error}".lower()
    data_keywords = ["anomaly score", "score_column", "score_null", "invalid_score", "missing required", "missing column"]
    api_keywords = ["timeout", "connection", "proxy", "407", "429", "500", "502", "503", "504", "http", "ssl", "tls", "network", "datarobot", "deployment", "service unavailable"]
    if any(k in error_text for k in data_keywords): return "ERROR_DATA", "PREDICT_DATA_ERROR"
    if any(k in error_text for k in api_keywords): return "ERROR_API", "PREDICT_API_ERROR"
    return "ERROR_SYSTEM", "PREDICT_SYSTEM_ERROR"

def write_success_status(pred_df, model_key):
    deployment_key = DEPLOYMENT_CONFIG[model_key]["deployment_key"]
    rows = []
    for _, row in pred_df.iterrows():
        rows.append({
            "Machine_No": row.get("Machine_No"), "Product_No_TBM": row.get("Product_No_TBM"),
            "Production_DateTime": row.get("Production_DateTime"), "Serial_No_TBM": row.get("Serial_No_TBM"),
            "status": row.get("judgment", "SUCCESS"), "anomaly_score": row.get("anomaly_score"),
            "judgment": row.get("judgment"), "trigger_rule": row.get("trigger_rule"),
            "deployment_key": deployment_key, "route_key": row.get("ROUTE_KEY"),
            "top_explanation": row.get("top_explanation"),
            "message": f"deployment={deployment_key}, score={row.get('anomaly_score')}, trigger={row.get('trigger_rule')}",
        })
    insert_status_rows(rows)

def write_error_status(df_part, model_key, error_message, error_status="ERROR_SYSTEM", trigger_rule="PREDICT_SYSTEM_ERROR"):
    deployment_info = DEPLOYMENT_CONFIG.get(model_key, {})
    deployment_key = deployment_info.get("deployment_key", str(model_key))
    rows = []
    for _, row in df_part.iterrows():
        rows.append({
            "Machine_No": row.get("Machine_No"), "Product_No_TBM": row.get("Product_No_TBM"),
            "Production_DateTime": row.get("Production_DateTime"), "Serial_No_TBM": row.get("Serial_No_TBM"),
            "status": error_status, "anomaly_score": None, "judgment": error_status,
            "trigger_rule": trigger_rule, "deployment_key": deployment_key,
            "route_key": row.get("ROUTE_KEY"), "top_explanation": None,
            "message": f"error={error_message}, status={error_status}",
        })
    insert_status_rows(rows)

def save_predictions(all_preds, path=None):
    if not all_preds: return
    out_df = pd.concat(all_preds, ignore_index=True)
    if path is None:
        os.makedirs(config.ANOMALY_SAVE_DIR, exist_ok=True)
        path = os.path.join(config.ANOMALY_SAVE_DIR, f"anomaly_predictions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
    out_df.to_csv(path, index=False, encoding="utf-8-sig")

def call_datarobot_genai_diagnosis(machine_no, route_key, trigger_rule, top_explanation):
    if not GEN_AI_URL: return None
    try:
        import requests
        prompt = f"你是轮胎成型机首席故障分析师。机台{machine_no},路径{route_key},触发{trigger_rule},主特征{top_explanation or 'N/A'}。请给出100字以内专业排查建议，不要客套话。"
        headers = {"Authorization": f"Bearer {GEN_AI_TOKEN}", "Content-Type": "application/json"}
        resp = requests.post(GEN_AI_URL, json=[{"prompt": prompt}], headers=headers, timeout=30)
        if resp.status_code != 200: return None
        data = resp.json()
        if isinstance(data, dict) and "data" in data and isinstance(data["data"], list) and len(data["data"]) > 0:
            return str(data["data"][0].get("prediction", "")).strip()[:200]
        return str(data).strip()[:200]
    except Exception as e:
        log_message(f"[GENAI] call failed: {e}")
        return None

print("[PREDICT+STATUS+GENAI] loaded")

# ============ 7.5 Explanation Trend Graph ============
def generate_explanation_trend_graph(machine_no, route_key, event_time, top_feature, trigger_rule=""):
    import matplotlib; matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    from datetime import timedelta
    try:
        if isinstance(event_time, str):
            event_dt = pd.to_datetime(event_time, errors="coerce")
        else:
            event_dt = pd.to_datetime(event_time)
        if pd.isna(event_dt): return
        win_before = EXPLANATION_TREND_WINDOW_MINUTES_BEFORE
        win_after = EXPLANATION_TREND_WINDOW_MINUTES_AFTER
        window_start = event_dt - timedelta(minutes=win_before)
        window_end = event_dt + timedelta(minutes=win_after)
        score_sql = f"SELECT Production_DateTime, anomaly_score FROM {STATUS_TABLE} WHERE Machine_No = ? AND route_key = ? AND Production_DateTime >= ? AND Production_DateTime <= ? AND anomaly_score IS NOT NULL ORDER BY Production_DateTime ASC"
        with get_connection() as conn:
            score_df = pd.read_sql(score_sql, conn, params=[machine_no, route_key, window_start, window_end])
        if score_df.empty: return
        score_df["_dt"] = pd.to_datetime(score_df["Production_DateTime"], errors="coerce")
        score_df["_score"] = pd.to_numeric(score_df["anomaly_score"], errors="coerce")
        score_df = score_df.dropna(subset=["_dt", "_score"])
        if score_df.empty: return
        output_dir = os.path.join(config.ANOMALY_SAVE_DIR, EXPLANATION_TREND_OUTPUT_SUBDIR)
        os.makedirs(output_dir, exist_ok=True)
        ts_str = event_dt.strftime("%Y%m%d_%H%M%S")
        safe_machine = str(machine_no).replace("/", "_").replace("\\\\", "_")
        safe_route = str(route_key).replace("/", "_").replace("\\\\", "_")
        output_path = os.path.join(output_dir, f"Trend_{safe_machine}_{safe_route}_{ts_str}.png")
        fig, ax_left = plt.subplots(figsize=(14, 6))
        ax_left.plot(score_df["_dt"], score_df["_score"], color="#1f77b4", linewidth=1.2, marker=".", markersize=3, label="anomaly_score", zorder=3)
        ax_left.set_ylabel("Anomaly Score", color="#1f77b4", fontsize=10)
        ax_left.set_ylim(0.0, 1.0)
        ax_left.tick_params(axis="y", labelcolor="#1f77b4")
        ax_left.axhline(y=ANOMALY_UCL, color="red", linestyle="--", linewidth=1.0, label=f"UCL={ANOMALY_UCL}", zorder=2)
        ax_left.axhline(y=ANOMALY_UWL, color="orange", linestyle="--", linewidth=1.0, label=f"UWL={ANOMALY_UWL}", zorder=2)
        ax_left.axvline(x=event_dt, color="magenta", linestyle=":", linewidth=1.0, label="Event Time", zorder=2)
        ax_left.xaxis.set_major_formatter(mdates.DateFormatter("%m/%d %H:%M"))
        ax_left.xaxis.set_major_locator(mdates.AutoDateLocator())
        fig.autofmt_xdate(rotation=30, ha="right")
        ax_left.set_title(f"Machine_No={machine_no}  ROUTE_KEY={route_key}\nEvent: {ts_str}  Rule: {trigger_rule}", fontsize=10, fontweight="bold")
        ax_left.legend(loc="upper left", fontsize=8, ncol=2)
        plt.tight_layout()
        fig.savefig(output_path, dpi=150, bbox_inches="tight")
        plt.close(fig)
    except Exception as e:
        log_message(f"[TREND_GRAPH] non-blocking failure Machine_No={machine_no}: {e}")
        try: plt.close("all")
        except: pass

def generate_explanation_trend_graphs(all_preds):
    t0 = time.perf_counter()
    try:
        merged_df = pd.concat(all_preds, ignore_index=True)
        if "judgment" not in merged_df.columns: return
        target_statuses = list(EXPLANATION_TREND_TARGET_STATUSES)
        event_df = merged_df[merged_df["judgment"].isin(target_statuses)].copy()
        if event_df.empty: return
        if "anomaly_score" in event_df.columns:
            event_df["_sort_score"] = pd.to_numeric(event_df["anomaly_score"], errors="coerce")
            event_df = event_df.sort_values("_sort_score", ascending=False)
        event_df = event_df.head(EXPLANATION_TREND_MAX_EVENTS)
        for _, row in event_df.iterrows():
            generate_explanation_trend_graph(
                machine_no=row.get("Machine_No"), route_key=row.get("ROUTE_KEY"),
                event_time=row.get("Production_DateTime"), top_feature=row.get("top_explanation"),
                trigger_rule=str(row.get("trigger_rule", "")),
            )
        log_elapsed(f"[TREND_GRAPH] generated {len(event_df)} graphs", t0)
    except Exception as e:
        log_message(f"[TREND_GRAPH] batch failed: {e}")

        last_detail_row = data_row - 1

        # ==== BLOCK 3: Scatter Charts ====
        import matplotlib; matplotlib.use("Agg"); import matplotlib.pyplot as plt
        import matplotlib.dates as mdates; from datetime import timedelta
        chart_row = last_detail_row + 4

        # Route-isolated time axis
        current_route_rows = abnormal_df[abnormal_df["ROUTE_KEY"].astype(str) == str(route_key)]
        if len(current_route_rows) > 0:
            all_pdt = pd.to_datetime(current_route_rows["Production_DateTime"], errors="coerce").dropna()
            ref_mno = current_route_rows.iloc[0].get("Machine_No")
        else:
            all_pdt = pd.Series()
            ref_mno = abnormal_df.iloc[0].get("Machine_No")
        if len(all_pdt) > 0:
            global_tmin = all_pdt.min(); global_tmax = all_pdt.max()
        else:
            global_tmin = ts_dt; global_tmax = ts_dt
        win_start = global_tmin - timedelta(minutes=win_before)
        win_end = global_tmax + timedelta(minutes=win_after)
        ref_rkey = route_key

        def _make_scatter_png(x_vals, y_vals, title_str, ylabel_str, ylim=None, hlines=None, colors=None):
            try:
                if colors is None: colors = [DARK_BLUE] * len(x_vals)
                fig, ax = plt.subplots(figsize=(8, 3))
                ax.scatter(x_vals, y_vals, c=colors, marker=".", s=12, alpha=0.85, zorder=3)
                ax.set_title(title_str, fontsize=9, fontweight="bold", color=DARK_BLUE)
                ax.set_ylabel(ylabel_str, fontsize=8, color=DARK_BLUE)
                ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M")); ax.tick_params(labelsize=7)
                if ylim: ax.set_ylim(*ylim)
                if hlines:
                    for yv, ls, lc, ll in hlines: ax.axhline(y=yv, linestyle=ls, color=lc, linewidth=0.8, label=ll)
                    ax.legend(fontsize=7, loc="upper right")
                fig.autofmt_xdate(rotation=20, ha="right"); plt.tight_layout()
                p = os.path.join(trend_dir, f"_scatter_{abs(hash(title_str))}.png")
                fig.savefig(p, dpi=150, bbox_inches="tight"); plt.close(fig); return p
            except Exception:
                try: plt.close("all")
                except: pass
                return None

        # Chart 1: Score Scatter
        chart1_ok = False
        try:
            with get_connection() as conn:
                sdf = pd.read_sql(f"SELECT Production_DateTime, anomaly_score FROM {STATUS_TABLE} WHERE Machine_No = ? AND route_key = ? AND Production_DateTime >= ? AND Production_DateTime <= ? AND anomaly_score IS NOT NULL ORDER BY Production_DateTime ASC", conn, params=[ref_mno, ref_rkey, win_start, win_end])
            if not sdf.empty:
                sdf["_dt"] = pd.to_datetime(sdf["Production_DateTime"]); sdf["_s"] = pd.to_numeric(sdf["anomaly_score"], errors="coerce")
                sdf = sdf.dropna(subset=["_dt", "_s"])
            if not sdf.empty:
                hl = [(ANOMALY_UCL, "--", "red", f"UCL={ANOMALY_UCL}"), (ANOMALY_UWL, "--", "orange", f"UWL={ANOMALY_UWL}")]
                png1 = _make_scatter_png(sdf["_dt"], sdf["_s"], f"Chart 1: Score Scatter | M={ref_mno} R={ref_rkey} | -{win_before}m/+{win_after}m", "Anomaly Score", ylim=(0.0, 1.0), hlines=hl, colors=[_score_to_color(sc) for sc in sdf["_s"]])
                if png1 and os.path.exists(png1):
                    ws.merge_cells(f"A{chart_row}:L{chart_row}")
                    ws.cell(row=chart_row, column=1, value=f"Chart 1: Anomaly Score Scatter (-{win_before}m/+{win_after}m)").font = Font(bold=True, size=10, color=DARK_BLUE)
                    chart_row += 1
                    for rr in range(chart_row, chart_row + 6): ws.row_dimensions[rr].height = 48
                    ws.merge_cells(f"A{chart_row}:L{chart_row + 5}")
                    img1 = Image(png1); img1.width = 960; img1.height = 360; ws.add_image(img1, f"A{chart_row}")
                    chart_row += 6; chart1_ok = True
        except Exception as e1:
            log_message(f"[EXCEL] Chart-1 failed: {e1}")
        if not chart1_ok:
            ws.merge_cells(f"A{chart_row}:L{chart_row}")
            ws.cell(row=chart_row, column=1, value="Chart 1: Score data unavailable.").font = Font(size=10, color="#999999", italic=True)
            chart_row += 1

        # Dynamic Charts: per-feature scatter
        unique_features = abnormal_df["top_explanation"].dropna().unique() if "top_explanation" in abnormal_df.columns else []
        chart_idx = 1
        for feat_name in unique_features:
            fname = str(feat_name).strip()
            if not fname: continue
            chart_idx += 1; fname_lower = fname.lower()
            feat_rows = abnormal_df[abnormal_df["top_explanation"].apply(lambda x: str(x).strip().lower() == fname_lower if pd.notna(x) else False)]
            if len(feat_rows) == 0:
                feat_rows = abnormal_df[abnormal_df["top_explanation"].apply(lambda x: str(x).strip().lower() == fname_lower.replace("-", "_") if pd.notna(x) else False)]
            try:
                if len(feat_rows) > 0:
                    current_feat_pdt = pd.to_datetime(feat_rows.iloc[0].get("Production_DateTime"), errors="coerce")
                    feat_mno = feat_rows.iloc[0].get("Machine_No")
                else:
                    ws.merge_cells(f"A{chart_row}:L{chart_row}")
                    ws.cell(row=chart_row, column=1, value=f"Chart {chart_idx}: {fname} -- Chart Data Unavailable").font = Font(size=10, color="#999999", italic=True)
                    chart_row += 1; continue
                feat_win_start = current_feat_pdt - timedelta(minutes=win_before)
                feat_win_end = current_feat_pdt + timedelta(minutes=win_after)
                feat_rkey = route_key

                # Step A: feature data
                col_candidates = list(dict.fromkeys([fname, fname.replace("-", "_"), fname_lower, fname_lower.replace("-", "_")]))
                fdf = pd.DataFrame()
                for ct in col_candidates:
                    try:
                        with get_connection() as conn:
                            fdf = pd.read_sql(f"SELECT Production_DateTime, [{ct}] AS _fv FROM {ANOMALY_HISTORICAL_FEATURE_VIEW} WHERE Machine_No = ? AND ROUTE_KEY = ? AND Production_DateTime >= ? AND Production_DateTime <= ? ORDER BY Production_DateTime ASC", conn, params=[feat_mno, feat_rkey, feat_win_start, feat_win_end])
                        if not fdf.empty and "_fv" in fdf.columns: break
                    except: continue
                if fdf.empty or "_fv" not in fdf.columns:
                    ws.merge_cells(f"A{chart_row}:L{chart_row}")
                    ws.cell(row=chart_row, column=1, value=f"Chart {chart_idx}: {fname} -- Chart Data Unavailable").font = Font(size=10, color="#999999", italic=True)
                    chart_row += 1; continue
                fdf["_dt"] = pd.to_datetime(fdf["Production_DateTime"]); fdf["_fv"] = pd.to_numeric(fdf["_fv"], errors="coerce")
                fdf = fdf.dropna(subset=["_dt", "_fv"])
                if len(fdf) == 0:
                    ws.merge_cells(f"A{chart_row}:L{chart_row}")
                    ws.cell(row=chart_row, column=1, value=f"Chart {chart_idx}: {fname} -- Chart Data Unavailable").font = Font(size=10, color="#999999", italic=True)
                    chart_row += 1; continue

                # Step B: score lookup (second-of-day matching)
                scoredf = pd.DataFrame()
                try:
                    with get_connection() as conn:
                        scoredf = pd.read_sql(f"SELECT Production_DateTime, anomaly_score FROM {STATUS_TABLE} WHERE Machine_No = ? AND route_key = ? AND Production_DateTime >= ? AND Production_DateTime <= ? AND anomaly_score IS NOT NULL", conn, params=[feat_mno, feat_rkey, feat_win_start, feat_win_end])
                    if not scoredf.empty:
                        scoredf["_dt"] = pd.to_datetime(scoredf["Production_DateTime"]); scoredf["_sc"] = pd.to_numeric(scoredf["anomaly_score"], errors="coerce")
                        scoredf = scoredf.dropna(subset=["_dt", "_sc"])
                except: pass

                # Second-of-day dictionary
                score_map = {}
                for _, sr in scoredf.iterrows():
                    sts = sr["_dt"].hour * 3600 + sr["_dt"].minute * 60 + sr["_dt"].second
                    sc = float(sr["_sc"])
                    if sc >= ANOMALY_UCL: score_map[sts] = "ANOMALY"
                    elif sc >= ANOMALY_UWL: score_map[sts] = "WARNING"
                    else: score_map[sts] = "NORMAL"

                # Color routing: NO ref_j -- unmatched points are pure DARK_BLUE
                feat_colors = []
                for _, frow in fdf.iterrows():
                    f_ts = frow["_dt"].hour * 3600 + frow["_dt"].minute * 60 + frow["_dt"].second
                    matched_status = "NORMAL"
                    for s_ts, status in score_map.items():
                        if abs(s_ts - f_ts) <= 45:
                            matched_status = status
                            break
                    if matched_status == "ANOMALY": feat_colors.append("#FF0000")
                    elif matched_status in ("TREND_ANOMALY", "WARNING"): feat_colors.append("#FF9900")
                    else: feat_colors.append(DARK_BLUE)

                png_f = _make_scatter_png(fdf["_dt"], fdf["_fv"], f"Chart {chart_idx}: {fname} Raw Value | M={feat_mno} R={feat_rkey} | -{win_before}m/+{win_after}m", f"{fname} (raw)", colors=feat_colors)
                if png_f and os.path.exists(png_f):
                    ws.merge_cells(f"A{chart_row}:L{chart_row}")
                    ws.cell(row=chart_row, column=1, value=f"Chart {chart_idx}: {fname} Raw Feature Scatter").font = Font(bold=True, size=10, color=DARK_BLUE)
                    chart_row += 1
                    for rr in range(chart_row, chart_row + 6): ws.row_dimensions[rr].height = 48
                    ws.merge_cells(f"A{chart_row}:L{chart_row + 5}")
                    img_f = Image(png_f); img_f.width = 960; img_f.height = 360; ws.add_image(img_f, f"A{chart_row}")
                    chart_row += 6
                else:
                    ws.merge_cells(f"A{chart_row}:L{chart_row}")
                    ws.cell(row=chart_row, column=1, value=f"Chart {chart_idx}: {fname} -- Chart Data Unavailable").font = Font(size=10, color="#999999", italic=True)
                    chart_row += 1
            except Exception as ef:
                log_message(f"[EXCEL] Chart-{chart_idx} failed for {fname}: {ef}")
                ws.merge_cells(f"A{chart_row}:L{chart_row}")
                ws.cell(row=chart_row, column=1, value=f"Chart {chart_idx}: {fname} -- Chart Data Unavailable").font = Font(size=10, color="#999999", italic=True)
                chart_row += 1
                try: plt.close("all")
                except: pass

        # Column width hard-lock
        ws.column_dimensions["A"].width = 17
        ws.column_dimensions["B"].width = 23
        ws.page_setup.orientation = "landscape"; ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
        wb.save(output_path)
        log_message(f"[EXCEL] dashboard saved path={output_path} total={len(merged_df)} abnormal={len(abnormal_df)} charts={chart_idx} win={win_before}m/+{win_after}m")
        return output_path

    except Exception as e:
        log_message(f"[EXCEL] report generation failed (non-blocking): {e}")
        log_message(traceback.format_exc())
        try:
            import matplotlib.pyplot as _plt; _plt.close("all")
        except: pass
        return None


# ============ 8. Main ============
def run_anomaly_detect():
    total_t0 = time.perf_counter()
    log_message("=" * 60)
    log_message("[TOTAL] start AnomalyDetect_new.py")
    raw_df = fetch_pending_data()
    if raw_df.empty:
        log_elapsed("[TOTAL] pending data = 0", total_t0)
        return 0
    routed = split_by_routing(raw_df)
    if not routed:
        log_elapsed("[TOTAL] valid routed data = 0", total_t0)
        return 0
    all_preds = []
    max_workers = getattr(config_anomaly, "MAX_WORKERS", 1)
    def run_one_batch(model_key, batch_df):
        pred_df = call_datarobot_api(model_key, batch_df)
        return model_key, batch_df, pred_df
    futures = {}
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        for model_key, df_part in routed.items():
            for start in range(0, len(df_part), BATCH_SIZE):
                batch_df = df_part.iloc[start:start + BATCH_SIZE].copy()
                futures[executor.submit(run_one_batch, model_key, batch_df)] = (model_key, batch_df)
        for future in as_completed(futures):
            model_key, batch_df = futures[future]
            try:
                model_key, batch_df, pred_df = future.result()
                pred_df = apply_rule2_consecutive_uwl(pred_df)
                all_preds.append(pred_df)
                write_success_status(pred_df, model_key)
            except Exception as e:
                err = str(e)
                error_status, trigger_rule = classify_prediction_exception(e)
                write_error_status(batch_df, model_key, err, error_status=error_status, trigger_rule=trigger_rule)
                log_message(f"[ERROR] {err}")
    if not all_preds:
        log_message("[SILENT] no predictions. skip csv.")
    else:
        merged_pred_df = pd.concat(all_preds, ignore_index=True)
        if "judgment" in merged_pred_df.columns:
            warning_count = int((merged_pred_df["judgment"] == "WARNING").sum())
            anomaly_count = int((merged_pred_df["judgment"] == "ANOMALY").sum())
            trend_count = int((merged_pred_df["judgment"] == "TREND_ANOMALY").sum())
            if warning_count == 0 and anomaly_count == 0 and trend_count == 0:
                log_message("[SILENT] All NORMAL. skip csv.")
            else:
                save_predictions(all_preds)
    # Trend graphs
    if ENABLE_EXPLANATION_TREND_GRAPH and all_preds:
        try: generate_explanation_trend_graphs(all_preds)
        except Exception as e: log_message(f"[TREND_GRAPH] error: {e}")
    # Excel reports -- one per route
    if ENABLE_EXCEL_REPORT and all_preds:
        excel_t0 = time.perf_counter()
        merged_all = pd.concat(all_preds, ignore_index=True)
        route_groups = {}
        if "ROUTE_KEY" in merged_all.columns:
            for rk, grp in merged_all.groupby("ROUTE_KEY"):
                route_groups[str(rk)] = [grp]
        else:
            route_groups["ALL"] = all_preds
        reports_generated = 0
        for route_key_str, route_preds in route_groups.items():
            try:
                report_path = generate_excel_report(route_preds, route_key=route_key_str)
                if report_path: reports_generated += 1
            except Exception as e:
                log_message(f"[EXCEL] report failed for {route_key_str}: {e}")
        log_elapsed(f"[EXCEL] reports generated={reports_generated}/{len(route_groups)}", excel_t0)
    log_elapsed(f"[TOTAL] finish", total_t0)
    log_message("=" * 60)
    return len(all_preds)

if __name__ == "__main__":
    run_anomaly_detect()

# ============ 7.6 Excel Dashboard Report ============
def generate_excel_report(all_preds, route_key="ALL"):
    """TireMonitor Dashboard (007) v4.x"""
    if not all_preds:
        log_message("[EXCEL] no predictions. skip.")
        return None
    try:
        merged_df = pd.concat(all_preds, ignore_index=True)
        if "judgment" not in merged_df.columns:
            log_message("[EXCEL] no judgment column. skip.")
            return None
        non_normal_mask = merged_df["judgment"] != "NORMAL"
        if not non_normal_mask.any():
            log_message(f"[EXCEL] all {len(merged_df)} NORMAL. skip (Silent Mode).")
            return None
        abnormal_df = merged_df[non_normal_mask].copy()
        if "anomaly_score" in abnormal_df.columns:
            abnormal_df["_sort_score"] = pd.to_numeric(abnormal_df["anomaly_score"], errors="coerce")
            abnormal_df = abnormal_df.sort_values("_sort_score", ascending=False)
        abnormal_df = abnormal_df.reset_index(drop=True)
        win_before = EXPLANATION_TREND_WINDOW_MINUTES_BEFORE
        win_after = EXPLANATION_TREND_WINDOW_MINUTES_AFTER
        os.makedirs(config.ANOMALY_SAVE_DIR, exist_ok=True)
        ts_dt = datetime.now()
        ts = ts_dt.strftime("%Y%m%d_%H%M%S")
        ts_display = ts_dt.strftime("%Y-%m-%d %H:%M:%S")
        output_path = os.path.join(config.ANOMALY_SAVE_DIR, f"Anomaly_Report_{route_key}_{ts}.xlsx")

        from openpyxl import Workbook
        from openpyxl.drawing.image import Image
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        wb = Workbook(); ws = wb.active; ws.title = "Dashboard_007"

        DARK_BLUE = "#1F4E78"; LIGHT_GRAY = "#F2F2F2"; WHITE = "#FFFFFF"
        header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
        header_font = Font(bold=True, size=11, color=WHITE, name="Calibri")
        title_font = Font(bold=True, size=16, color=WHITE, name="Calibri")
        label_font = Font(bold=True, size=10, color=DARK_BLUE, name="Calibri")
        value_font = Font(size=10, color="#333333", name="Calibri")
        data_font = Font(size=10, color="#333333", name="Calibri")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        label_align = Alignment(horizontal="left", vertical="center")
        value_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        data_align = Alignment(vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style="thin", color=DARK_BLUE), right=Side(style="thin", color=DARK_BLUE), top=Side(style="thin", color=DARK_BLUE), bottom=Side(style="thin", color=DARK_BLUE))
        trend_dir = os.path.join(config.ANOMALY_SAVE_DIR, EXPLANATION_TREND_OUTPUT_SUBDIR)
        os.makedirs(trend_dir, exist_ok=True)

        def _score_to_color(sc):
            if pd.isna(sc): return DARK_BLUE
            if float(sc) >= ANOMALY_UCL: return "#FF0000"
            if float(sc) >= ANOMALY_UWL: return "#FF9900"
            return DARK_BLUE

        # BLOCK 1 (Row 1-6)
        ws.column_dimensions["A"].width = 22; ws.column_dimensions["B"].width = 32
        ws.merge_cells("A1:L1")
        c = ws.cell(row=1, column=1, value="TireMonitor Anomaly Detection Dashboard (007)")
        c.font = title_font; c.fill = header_fill; c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 36
        for rnum, label, val in [(2, "Target System", "AnomalyDetect_new (007)"), (3, "Batch Timestamp", ts_display), (4, "Total Inspected", len(merged_df)), (5, "Abnormal Detected", non_normal_mask.sum())]:
            ws.row_dimensions[rnum].height = 20
            ca = ws.cell(row=rnum, column=1, value=label); ca.font = label_font; ca.alignment = label_align
            ca.fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid"); ca.border = thin_border
            cb = ws.cell(row=rnum, column=2, value=val); cb.font = value_font; cb.alignment = value_align; cb.border = thin_border
        if "trigger_rule" in abnormal_df.columns:
            rc = abnormal_df["trigger_rule"].value_counts()
            rule_summary = ", ".join([f"{r}({c})" for r, c in rc.items()])
        else:
            rule_summary = "N/A"
        ws.row_dimensions[6].height = 20
        ca6 = ws.cell(row=6, column=1, value="Alarm Trigger Rule"); ca6.font = label_font; ca6.alignment = label_align
        ca6.fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid"); ca6.border = thin_border
        cb6 = ws.cell(row=6, column=2, value=rule_summary); cb6.font = value_font; cb6.alignment = value_align; cb6.border = thin_border
        ws.row_dimensions[7].height = 6

        # BLOCK 2 (Row 10+) Detail Table A-L
        detail_start = 10
        detail_headers = ["Rank", "Machine_No", "Production_DateTime", "Product_No_TBM", "Serial_No_TBM", "ROUTE_KEY", "anomaly_score", "judgment", "trigger_rule", "top_explanation", "Top_Explanation_Value", "AI_Diagnosis_Comment"]
        NUM_COLS = len(detail_headers); AI_COL = NUM_COLS
        ws.row_dimensions[detail_start].height = 22
        for ci, hdr in enumerate(detail_headers, 1):
            c = ws.cell(row=detail_start, column=ci, value=hdr); c.font = header_font; c.fill = header_fill; c.alignment = header_align; c.border = thin_border
        for ci, w in enumerate([6, 14, 22, 18, 18, 18, 16, 14, 22, 30, 22, 40], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        ai_cache = {}; ai_barcodes = set()
        if ENABLE_AI_DIAGNOSIS:
            for _, row in abnormal_df.iterrows():
                judgment = str(row.get("judgment", ""))
                if judgment != "NORMAL":
                    try:
                        barcode = "".join(str(row.get("Serial_No_TBM")).split())
                        if not barcode or barcode in ai_barcodes: continue
                        ai_barcodes.add(barcode)
                        comment = call_datarobot_genai_diagnosis(
                            machine_no=row.get("Machine_No"), route_key=row.get("ROUTE_KEY"),
                            trigger_rule=str(row.get("trigger_rule", "")),
                            top_explanation=str(row.get("top_explanation", "") if pd.notna(row.get("top_explanation")) else "N/A"),
                        )
                        if comment: ai_cache[barcode] = str(comment)[:500]
                    except Exception as diag_err:
                        log_message(f"[GENAI] pre-fetch failed: {diag_err}")

        data_row = detail_start + 1
        for _, row in abnormal_df.iterrows():
            rank = list(abnormal_df.index).index(row.name) + 1
            ws.row_dimensions[data_row].height = 20
            judgment = str(row.get("judgment", "")); mno = row.get("Machine_No"); rkey = row.get("ROUTE_KEY")
            pdt = row.get("Production_DateTime"); tfeat = row.get("top_explanation")
            pdt_str = pdt.strftime("%Y-%m-%d %H:%M:%S") if isinstance(pdt, pd.Timestamp) else str(pdt)
            score_val = round(float(row.get("anomaly_score", 0)), 4) if pd.notna(row.get("anomaly_score")) else None
            vals = [rank, mno, pdt_str, row.get("Product_No_TBM"), row.get("Serial_No_TBM"), rkey, score_val, judgment, str(row.get("trigger_rule", "")), str(tfeat) if pd.notna(tfeat) else None, None, None]
            for ci, v in enumerate(vals, 1):
                if v is not None or ci >= 11:
                    c = ws.cell(row=data_row, column=ci, value=v if ci < 11 else None); c.font = data_font; c.alignment = data_align; c.border = thin_border
            current_barcode = "".join(str(row.get("Serial_No_TBM")).split())
            cached_comment = ai_cache.get(current_barcode)
            if cached_comment:
                c12 = ws.cell(row=data_row, column=AI_COL, value=cached_comment)
                c12.font = Font(size=9, color="#333333", name="Calibri"); c12.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True); c12.border = thin_border
            if judgment == "ANOMALY": rf = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif judgment == "TREND_ANOMALY": rf = PatternFill(start_color="FFEBAF", end_color="FFEBAF", fill_type="solid")
            elif judgment == "WARNING": rf = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            else: rf = None
            if rf:
                for ci in range(1, NUM_COLS + 1): ws.cell(row=data_row, column=ci).fill = rf
            data_row += 1
        last_detail_row = data_row - 1

        # BLOCK 3: Scatter Charts
        import matplotlib; matplotlib.use("Agg"); import matplotlib.pyplot as plt
        import matplotlib.dates as mdates; from datetime import timedelta
        chart_row = last_detail_row + 4

        current_route_rows = abnormal_df[abnormal_df["ROUTE_KEY"].astype(str) == str(route_key)]
        if len(current_route_rows) > 0:
            all_pdt = pd.to_datetime(current_route_rows["Production_DateTime"], errors="coerce").dropna()
            ref_mno = current_route_rows.iloc[0].get("Machine_No")
        else:
            all_pdt = pd.Series()
            ref_mno = abnormal_df.iloc[0].get("Machine_No")
        if len(all_pdt) > 0:
            global_tmin = all_pdt.min(); global_tmax = all_pdt.max()
        else:
            global_tmin = ts_dt; global_tmax = ts_dt
        win_start = global_tmin - timedelta(minutes=win_before)
        win_end = global_tmax + timedelta(minutes=win_after)
        ref_rkey = route_key

        def _make_scatter_png(x_vals, y_vals, title_str, ylabel_str, ylim=None, hlines=None, colors=None):
            try:
                if colors is None: colors = [DARK_BLUE] * len(x_vals)
                fig, ax = plt.subplots(figsize=(8, 3))
                ax.scatter(x_vals, y_vals, c=colors, marker=".", s=12, alpha=0.85, zorder=3)
                ax.set_title(title_str, fontsize=9, fontweight="bold", color=DARK_BLUE)
                ax.set_ylabel(ylabel_str, fontsize=8, color=DARK_BLUE)
                ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M")); ax.tick_params(labelsize=7)
                if ylim: ax.set_ylim(*ylim)
                if hlines:
                    for yv, ls, lc, ll in hlines: ax.axhline(y=yv, linestyle=ls, color=lc, linewidth=0.8, label=ll)
                    ax.legend(fontsize=7, loc="upper right")
                fig.autofmt_xdate(rotation=20, ha="right"); plt.tight_layout()
                p = os.path.join(trend_dir, f"_scatter_{abs(hash(title_str))}.png")
                fig.savefig(p, dpi=150, bbox_inches="tight"); plt.close(fig); return p
            except Exception:
                try: plt.close("all")
                except: pass
                return None

        # Chart 1
        chart1_ok = False
        try:
            with get_connection() as conn:
                sdf = pd.read_sql(f"SELECT Production_DateTime, anomaly_score FROM {STATUS_TABLE} WHERE Machine_No = ? AND route_key = ? AND Production_DateTime >= ? AND Production_DateTime <= ? AND anomaly_score IS NOT NULL ORDER BY Production_DateTime ASC", conn, params=[ref_mno, ref_rkey, win_start, win_end])
            if not sdf.empty:
                sdf["_dt"] = pd.to_datetime(sdf["Production_DateTime"]); sdf["_s"] = pd.to_numeric(sdf["anomaly_score"], errors="coerce")
                sdf = sdf.dropna(subset=["_dt", "_s"])
            if not sdf.empty:
                hl = [(ANOMALY_UCL, "--", "red", f"UCL={ANOMALY_UCL}"), (ANOMALY_UWL, "--", "orange", f"UWL={ANOMALY_UWL}")]
                png1 = _make_scatter_png(sdf["_dt"], sdf["_s"], f"Chart 1: Score Scatter | M={ref_mno} R={ref_rkey} | -{win_before}m/+{win_after}m", "Anomaly Score", ylim=(0.0, 1.0), hlines=hl, colors=[_score_to_color(sc) for sc in sdf["_s"]])
                if png1 and os.path.exists(png1):
                    ws.merge_cells(f"A{chart_row}:L{chart_row}")
                    ws.cell(row=chart_row, column=1, value=f"Chart 1: Anomaly Score Scatter (-{win_before}m/+{win_after}m)").font = Font(bold=True, size=10, color=DARK_BLUE)
                    chart_row += 1
                    for rr in range(chart_row, chart_row + 6): ws.row_dimensions[rr].height = 48
                    ws.merge_cells(f"A{chart_row}:L{chart_row + 5}")
                    img1 = Image(png1); img1.width = 960; img1.height = 360; ws.add_image(img1, f"A{chart_row}")
                    chart_row += 6; chart1_ok = True
        except Exception as e1:
            log_message(f"[EXCEL] Chart-1 failed: {e1}")
        if not chart1_ok:
            ws.merge_cells(f"A{chart_row}:L{chart_row}")
            ws.cell(row=chart_row, column=1, value="Chart 1: Score data unavailable.").font = Font(size=10, color="#999999", italic=True)
            chart_row += 1

        # Dynamic feature charts
        unique_features = abnormal_df["top_explanation"].dropna().unique() if "top_explanation" in abnormal_df.columns else []
        chart_idx = 1
        for feat_name in unique_features:
            fname = str(feat_name).strip()
            if not fname: continue
            chart_idx += 1; fname_lower = fname.lower()
            feat_rows = abnormal_df[abnormal_df["top_explanation"].apply(lambda x: str(x).strip().lower() == fname_lower if pd.notna(x) else False)]
            if len(feat_rows) == 0:
                feat_rows = abnormal_df[abnormal_df["top_explanation"].apply(lambda x: str(x).strip().lower() == fname_lower.replace("-", "_") if pd.notna(x) else False)]
            try:
                if len(feat_rows) > 0:
                    current_feat_pdt = pd.to_datetime(feat_rows.iloc[0].get("Production_DateTime"), errors="coerce")
                    feat_mno = feat_rows.iloc[0].get("Machine_No")
                else:
                    ws.merge_cells(f"A{chart_row}:L{chart_row}")
                    ws.cell(row=chart_row, column=1, value=f"Chart {chart_idx}: {fname} -- Chart Data Unavailable").font = Font(size=10, color="#999999", italic=True)
                    chart_row += 1; continue
                feat_win_start = current_feat_pdt - timedelta(minutes=win_before)
                feat_win_end = current_feat_pdt + timedelta(minutes=win_after)
                feat_rkey = route_key

                col_candidates = list(dict.fromkeys([fname, fname.replace("-", "_"), fname_lower, fname_lower.replace("-", "_")]))
                fdf = pd.DataFrame()
                for ct in col_candidates:
                    try:
                        with get_connection() as conn:
                            fdf = pd.read_sql(f"SELECT Production_DateTime, [{ct}] AS _fv FROM {ANOMALY_HISTORICAL_FEATURE_VIEW} WHERE Machine_No = ? AND ROUTE_KEY = ? AND Production_DateTime >= ? AND Production_DateTime <= ? ORDER BY Production_DateTime ASC", conn, params=[feat_mno, feat_rkey, feat_win_start, feat_win_end])
                        if not fdf.empty and "_fv" in fdf.columns: break
                    except: continue
                if fdf.empty or "_fv" not in fdf.columns:
                    ws.merge_cells(f"A{chart_row}:L{chart_row}")
                    ws.cell(row=chart_row, column=1, value=f"Chart {chart_idx}: {fname} -- Chart Data Unavailable").font = Font(size=10, color="#999999", italic=True)
                    chart_row += 1; continue
                fdf["_dt"] = pd.to_datetime(fdf["Production_DateTime"]); fdf["_fv"] = pd.to_numeric(fdf["_fv"], errors="coerce")
                fdf = fdf.dropna(subset=["_dt", "_fv"])
                if len(fdf) == 0:
                    ws.merge_cells(f"A{chart_row}:L{chart_row}")
                    ws.cell(row=chart_row, column=1, value=f"Chart {chart_idx}: {fname} -- Chart Data Unavailable").font = Font(size=10, color="#999999", italic=True)
                    chart_row += 1; continue

                scoredf = pd.DataFrame()
                try:
                    with get_connection() as conn:
                        scoredf = pd.read_sql(f"SELECT Production_DateTime, anomaly_score FROM {STATUS_TABLE} WHERE Machine_No = ? AND route_key = ? AND Production_DateTime >= ? AND Production_DateTime <= ? AND anomaly_score IS NOT NULL", conn, params=[feat_mno, feat_rkey, feat_win_start, feat_win_end])
                    if not scoredf.empty:
                        scoredf["_dt"] = pd.to_datetime(scoredf["Production_DateTime"]); scoredf["_sc"] = pd.to_numeric(scoredf["anomaly_score"], errors="coerce")
                        scoredf = scoredf.dropna(subset=["_dt", "_sc"])
                except: pass

                score_map = {}
                for _, sr in scoredf.iterrows():
                    sts = sr["_dt"].hour * 3600 + sr["_dt"].minute * 60 + sr["_dt"].second
                    sc = float(sr["_sc"])
                    if sc >= ANOMALY_UCL: score_map[sts] = "ANOMALY"
                    elif sc >= ANOMALY_UWL: score_map[sts] = "WARNING"
                    else: score_map[sts] = "NORMAL"

                feat_colors = []
                for _, frow in fdf.iterrows():
                    f_ts = frow["_dt"].hour * 3600 + frow["_dt"].minute * 60 + frow["_dt"].second
                    matched_status = "NORMAL"
                    for s_ts, status in score_map.items():
                        if abs(s_ts - f_ts) <= 45:
                            matched_status = status
                            break
                    if matched_status == "ANOMALY": feat_colors.append("#FF0000")
                    elif matched_status in ("TREND_ANOMALY", "WARNING"): feat_colors.append("#FF9900")
                    else: feat_colors.append(DARK_BLUE)

                png_f = _make_scatter_png(fdf["_dt"], fdf["_fv"], f"Chart {chart_idx}: {fname} Raw Value | M={feat_mno} R={feat_rkey} | -{win_before}m/+{win_after}m", f"{fname} (raw)", colors=feat_colors)
                if png_f and os.path.exists(png_f):
                    ws.merge_cells(f"A{chart_row}:L{chart_row}")
                    ws.cell(row=chart_row, column=1, value=f"Chart {chart_idx}: {fname} Raw Feature Scatter").font = Font(bold=True, size=10, color=DARK_BLUE)
                    chart_row += 1
                    for rr in range(chart_row, chart_row + 6): ws.row_dimensions[rr].height = 48
                    ws.merge_cells(f"A{chart_row}:L{chart_row + 5}")
                    img_f = Image(png_f); img_f.width = 960; img_f.height = 360; ws.add_image(img_f, f"A{chart_row}")
                    chart_row += 6
                else:
                    ws.merge_cells(f"A{chart_row}:L{chart_row}")
                    ws.cell(row=chart_row, column=1, value=f"Chart {chart_idx}: {fname} -- Chart Data Unavailable").font = Font(size=10, color="#999999", italic=True)
                    chart_row += 1
            except Exception as ef:
                log_message(f"[EXCEL] Chart-{chart_idx} failed for {fname}: {ef}")
                ws.merge_cells(f"A{chart_row}:L{chart_row}")
                ws.cell(row=chart_row, column=1, value=f"Chart {chart_idx}: {fname} -- Chart Data Unavailable").font = Font(size=10, color="#999999", italic=True)
                chart_row += 1
                try: plt.close("all")
                except: pass

        ws.column_dimensions["A"].width = 17
        ws.column_dimensions["B"].width = 23
        ws.page_setup.orientation = "landscape"; ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
        wb.save(output_path)
        log_message(f"[EXCEL] dashboard saved path={output_path} total={len(merged_df)} abnormal={len(abnormal_df)} charts={chart_idx} win={win_before}m/+{win_after}m")
        return output_path
    except Exception as e:
        log_message(f"[EXCEL] report generation failed (non-blocking): {e}")
        log_message(traceback.format_exc())
        try:
            import matplotlib.pyplot as _plt; _plt.close("all")
        except: pass
        return None


# ============ 8. Main ============
def run_anomaly_detect():
    total_t0 = time.perf_counter()
    log_message("=" * 60)
    log_message("[TOTAL] start AnomalyDetect_new.py")
    raw_df = fetch_pending_data()
    if raw_df.empty:
        log_elapsed("[TOTAL] pending data = 0", total_t0)
        return 0
    routed = split_by_routing(raw_df)
    if not routed:
        log_elapsed("[TOTAL] valid routed data = 0", total_t0)
        return 0
    all_preds = []
    max_workers = getattr(config_anomaly, "MAX_WORKERS", 1)
    def run_one_batch(model_key, batch_df):
        pred_df = call_datarobot_api(model_key, batch_df)
        return model_key, batch_df, pred_df
    futures = {}
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        for model_key, df_part in routed.items():
            for start in range(0, len(df_part), BATCH_SIZE):
                batch_df = df_part.iloc[start:start + BATCH_SIZE].copy()
                futures[executor.submit(run_one_batch, model_key, batch_df)] = (model_key, batch_df)
        for future in as_completed(futures):
            model_key, batch_df = futures[future]
            try:
                model_key, batch_df, pred_df = future.result()
                pred_df = apply_rule2_consecutive_uwl(pred_df)
                all_preds.append(pred_df)
                write_success_status(pred_df, model_key)
            except Exception as e:
                err = str(e)
                error_status, trigger_rule = classify_prediction_exception(e)
                write_error_status(batch_df, model_key, err, error_status=error_status, trigger_rule=trigger_rule)
                log_message(f"[ERROR] {err}")
    if not all_preds:
        log_message("[SILENT] no predictions. skip csv.")
    else:
        merged_pred_df = pd.concat(all_preds, ignore_index=True)
        if "judgment" in merged_pred_df.columns:
            warning_count = int((merged_pred_df["judgment"] == "WARNING").sum())
            anomaly_count = int((merged_pred_df["judgment"] == "ANOMALY").sum())
            trend_count = int((merged_pred_df["judgment"] == "TREND_ANOMALY").sum())
            if warning_count == 0 and anomaly_count == 0 and trend_count == 0:
                log_message("[SILENT] All NORMAL. skip csv.")
            else:
                save_predictions(all_preds)
    if ENABLE_EXPLANATION_TREND_GRAPH and all_preds:
        try: generate_explanation_trend_graphs(all_preds)
        except Exception as e: log_message(f"[TREND_GRAPH] error: {e}")
    if ENABLE_EXCEL_REPORT and all_preds:
        excel_t0 = time.perf_counter()
        merged_all = pd.concat(all_preds, ignore_index=True)
        route_groups = {}
        if "ROUTE_KEY" in merged_all.columns:
            for rk, grp in merged_all.groupby("ROUTE_KEY"):
                route_groups[str(rk)] = [grp]
        else:
            route_groups["ALL"] = all_preds
        reports_generated = 0
        for route_key_str, route_preds in route_groups.items():
            try:
                report_path = generate_excel_report(route_preds, route_key=route_key_str)
                if report_path: reports_generated += 1
            except Exception as e:
                log_message(f"[EXCEL] report failed for {route_key_str}: {e}")
        log_elapsed(f"[EXCEL] reports generated={reports_generated}/{len(route_groups)}", excel_t0)
    log_elapsed(f"[TOTAL] finish", total_t0)
    log_message("=" * 60)
    return len(all_preds)

if __name__ == "__main__":
    run_anomaly_detect()
